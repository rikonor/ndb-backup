import os, sys, json, time
import cloudstorage as gcs
from openpyxl import load_workbook

from BaseHandler import BaseHandler
from AuthHandlers import *
from models import models
from models.models_map import ModelsMap
from security import hashes

from google.appengine.ext import ndb
from google.appengine.ext import blobstore
from google.appengine.ext.webapp import blobstore_handlers

BUCKET = '/ndb-backup.appspot.com'

OBJ_START_ROW = 3

#--------------------------------------------------------------
# Recovery methods
#--------------------------------------------------------------
def getRows(ws):
	result = 0
	while(ws.cell(row=result+OBJ_START_ROW, column=0).value): result += 1
	return result

def createKey(type, id):
	return ndb.Key(type, id)

def processRow(ws, row, clsInfo):
	cls = getattr(models, clsInfo['repr'])
	column = 0
	# New with ID
	newId = int(ws.cell(row=row, column=column).value)
	newObj = cls(id=newId)
	column += 1
	# Properties
	for prop in clsInfo["properties"]:
		newProp = ws.cell(row=row, column=column).value
		if prop['type'] == 'text':   newProp = str(newProp)
		if prop['type'] == 'number': newProp = int(newProp)
		setattr(newObj, prop['name'], newProp)
		column += 1
	# References
	for ref in clsInfo["references"]:
		refValue = str(ws.cell(row=row, column=column).value)
		keys = refValue.split(",")
		if len(keys) == 1:
			newRef = createKey(ref['kind'], int(keys[0]))
		else:
			newRef = []
			for i in range(len(keys)):
				newRef.append(createKey(ref['kind'], int(keys[i])))
		setattr(newObj, ref['name'], newRef)
		column += 1
	# save
	newObj.put()

def processSheet(ws, clsInfo):
	row = OBJ_START_ROW
	rows = getRows(ws)
	for row in range(rows):
		processRow(ws, row+OBJ_START_ROW, clsInfo)

def startRecovery(wb):
	Map = ModelsMap.Map
	for clsInfo in Map:
		ws = wb.get_sheet_by_name(clsInfo['repr'])
		processSheet(ws, clsInfo)	
#--------------------------------------------------------------
# ndbRecover handler
#--------------------------------------------------------------
class ndbRecover(BaseHandler):

    def get(self):
        # user = Authenticate(self.request)
        # if not user:
        #     return self.redirect("/login")

        upload_url = blobstore.create_upload_url('/ndbrecoverupload')
        return self.render("ndbrecover.html", upload_url=upload_url)

class ndbRecoverUpload(blobstore_handlers.BlobstoreUploadHandler):

	def post(self):
		# user = Authenticate(self.request)
		# if not user:
		# 	return self.redirect("/login")
		blob_info = self.get_uploads('upload')[0]
		reader = blob_info.open()
		wb = load_workbook(reader)
		startRecovery(wb)
		blob_info.delete()
		self.redirect("/")
#--------------------------------------------------------------