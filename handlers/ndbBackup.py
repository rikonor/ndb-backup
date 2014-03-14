import os, sys, json, time
import cloudstorage as gcs
from openpyxl import Workbook

from BaseHandler import BaseHandler
from AuthHandlers import *
from models import models
from models.models_map import ModelsMap
from security import hashes

my_default_retry_params = gcs.RetryParams(initial_delay=0.2,
                                          max_delay=5.0,
                                          backoff_factor=2,
                                          max_retry_period=15)
gcs.set_default_retry_params(my_default_retry_params)

BUCKET = '/ndb-backup.appspot.com'

#--------------------------------------------------------------
# ndbBackupManager
#--------------------------------------------------------------
#   -- XLSX Methods --   #
#------------------------#
def initWorkbook():
    # return an empty workbook
    wb = Workbook()
    wb.remove_sheet(wb.active)
    return wb

def newSheet(wb, name):
    # return new sheet with name and class value
    ws = wb.create_sheet()
    ws.title = name
    ws.cell('A1').value = name
    return ws

def setSheetProperties(ws, clsInfo):
    # Set ID and Type
    ws.cell(row=1, column=0).value = "ID"
    ws.cell(row=2, column=0).value = "number"
    # Set Properties and Types
    row = 1; column = 1
    for prop in clsInfo["properties"]:
        ws.cell(row=row  , column=column).value = prop["name"]
        ws.cell(row=row+1, column=column).value = prop["type"]
        column += 1
    # Set References and Types
    for ref in clsInfo["references"]:
        ws.cell(row=row  , column=column).value = ref["name"]
        ws.cell(row=row+1, column=column).value = ref["kind"]
        column += 1

def putObjInWorksheet(ws, row, obj, clsInfo):
    column = 0
    # put ID
    ws.cell(row=row, column=column).value = obj.key.id()
    # put Properties
    column = 1
    for prop in clsInfo["properties"]:
        value = getattr(obj, prop["name"])
        ws.cell(row=row, column=column).value = value
        column += 1
    # put references
    for ref in clsInfo["references"]:
        keys = getattr(obj, ref["name"])
        if type(keys) == list:
            ws.cell(row=row, column=column).value = ",".join([str(key.id()) for key in keys])
        else:
            ws.cell(row=row, column=column).value = keys.id()
        column += 1

def saveWorkbook(wb):
    # set filename as "backup_*current-datetime*"
    FILENAME = BUCKET + "/backup_" + datetime.datetime.now().strftime("%m:%d:%Y-%T") + ".xlsx"
    # new cloud storage file
    gcs_file = gcs.open(FILENAME,
                        'w',
                        content_type='application/vnd.ms-excel',
                        retry_params=my_default_retry_params)
    # save workbook to gcs file
    wb.save(gcs_file)
    # finalize the gcs file
    gcs_file.close()

#------------------------#
#   -- NDB Methods --    #
#------------------------#

def getAllFromClass(name):
    objClass = getattr(models, name)
    return objClass.query().fetch()

#------------------------#

def startBackup():
    wb = initWorkbook()
    Map = ModelsMap.Map
    # workbook and models map ready
    # going over all the classes
    for clsInfo in Map:
        # new sheet per class
        ws = newSheet(wb, clsInfo["repr"])
        setSheetProperties(ws, clsInfo)
        objs = getAllFromClass(clsInfo["repr"])
        row = 3
        for obj in objs:
            putObjInWorksheet(ws, row, obj, clsInfo)
            row += 1

    saveWorkbook(wb)
#--------------------------------------------------------------
# ndbBackup handler
#--------------------------------------------------------------
class ndbBackup(BaseHandler):

    def get(self):
        # user = Authenticate(self.request)
        # if not user:
        #     return self.redirect("/login")
        #wb = Workbook()
        #ws = wb.active
        #ws.cell('A1').value = "Hello World!"

        # gcs_file = gcs.open(filename)
        # self.response.write(gcs_file.readline())
        # gcs_file.seek(-1024, os.SEEK_END)
        # self.response.write(gcs_file.read())
        # gcs_file.close()

        # stat = gcs.stat(filename)
        # self.response.write(repr(stat))

        # gcs.delete(filename)
        startBackup()

        return self.response.write("Done.")

#--------------------------------------------------------------